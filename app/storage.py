# storage.py
import os
import datetime
import sqlite3
import json
from openpyxl import Workbook, load_workbook
# 导入根目录的配置
import config


# === 原有的 ExcelSaver 保持不变 ===
class ExcelSaver:
    # 默认使用 config 中定义的路径
    def __init__(self, filename=config.EXCEL_PATH):
        self.filename = filename
        self.init_file()

    def init_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "账单记录"
            headers = ["记录时间", "截图文件名", "商户/商品", "分类", "金额", "备注(原始数据)"]
            ws.append(headers)
            wb.save(self.filename)
            print(f"📘 [Excel] 已创建新账本: {self.filename}")

    def save(self, data, image_name):
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            merchant_clean = data["merchant"].split("￥")[0].replace(">", "").strip()
            row = [
                now_time, image_name, merchant_clean,
                data["category"], data["amount"], str(data["raw_text"])[:50] + "..."
            ]
            ws.append(row)
            wb.save(self.filename)
            print(f"✅ [Excel] 写入成功")
        except PermissionError:
            print(f"❌ [Excel] 写入失败: 请先关闭打开的文件！")


# === ✨ 新增 DatabaseSaver ===
class DatabaseSaver:
    # 默认使用 config 中定义的路径
    def __init__(self, db_name=config.DB_PATH):
        self.db_name = db_name
        self.init_db()

    def init_db(self):
        """初始化数据库表结构"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        # 创建 bills 表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_time TEXT,
                image_name TEXT,
                merchant TEXT,
                category TEXT,
                amount REAL,
                raw_text TEXT
            )
        ''')
        conn.commit()
        conn.close()
        print(f"📘 [DB] 数据库连接就绪: {self.db_name}")

    def save(self, data, image_name):
        """插入一条新记录"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        merchant_clean = data["merchant"].split("￥")[0].replace(">", "").strip()

        # 将原始文本列表转为 JSON 字符串存储
        raw_text_json = json.dumps(data["raw_text"], ensure_ascii=False)

        cursor.execute('''
            INSERT INTO bills (record_time, image_name, merchant, category, amount, raw_text)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (now_time, image_name, merchant_clean, data["category"], data["amount"], raw_text_json))

        conn.commit()
        conn.close()
        print(f"✅ [DB] 数据已存入数据库")