# enhanced_storage.py - Enhanced storage system with new features
import os
import datetime
import sqlite3
import json
from typing import List, Dict, Optional, Any
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
import config


@dataclass
class EnhancedBill:
    """Enhanced bill data model"""
    id: Optional[int] = None
    ledger_id: Optional[int] = None
    filename: str = ""
    merchant: str = ""
    amount: float = 0.0
    category_id: Optional[int] = None
    category: str = ""
    bill_date: str = ""  # YYYY-MM-DD format
    created_at: str = ""
    updated_at: str = ""
    raw_text: List[str] = None
    is_manual: bool = False
    include_in_budget: bool = True
    
    def __post_init__(self):
        if self.raw_text is None:
            self.raw_text = []
        if not self.bill_date:
            self.bill_date = datetime.date.today().strftime("%Y-%m-%d")
        if not self.created_at:
            self.created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self.updated_at:
            self.updated_at = self.created_at
        if self.include_in_budget is None:
            self.include_in_budget = True


@dataclass
class CategoryRule:
    """Category rule data model"""
    id: Optional[int] = None
    keyword: str = ""
    category_id: Optional[int] = None
    category: str = ""
    ledger_id: Optional[int] = None
    priority: int = 2
    is_weak: bool = False
    created_at: str = ""
    updated_at: str = ""
    
    def __post_init__(self):
        if not self.created_at:
            self.created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self.updated_at:
            self.updated_at = self.created_at


@dataclass
class CategoryGroup:
    """Category group data model"""
    id: Optional[int] = None
    ledger_id: Optional[int] = None
    major: str = ""
    minor: str = ""
    created_at: str = ""
    updated_at: str = ""

    def __post_init__(self):
        if not self.created_at:
            self.created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self.updated_at:
            self.updated_at = self.created_at


@dataclass
class RecurringRule:
    """Recurring bill rule data model"""
    id: Optional[int] = None
    ledger_id: Optional[int] = None
    amount: float = 0.0
    keyword: str = ""
    category_id: Optional[int] = None
    category: str = ""
    note: str = ""
    schedule_type: str = "weekly"  # weekly | monthly
    schedule_value: List[int] = None
    start_date: str = ""  # YYYY-MM-DD
    end_date: Optional[str] = None
    enabled: bool = True
    include_in_budget: bool = True
    created_at: str = ""
    updated_at: str = ""

    def __post_init__(self):
        today = datetime.date.today().strftime("%Y-%m-%d")
        if not self.start_date:
            self.start_date = today
        if self.schedule_value is None:
            self.schedule_value = [1]
        if not self.created_at:
            self.created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not self.updated_at:
            self.updated_at = self.created_at
        if self.include_in_budget is None:
            self.include_in_budget = True

class EnhancedDatabaseManager:
    """Enhanced database manager with new schema and features"""
    
    def __init__(self, db_name=config.DB_PATH):
        self.db_name = db_name
        self.init_db()
    
    def init_db(self):
        """Initialize enhanced database schema"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # Ledgers table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ledgers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                monthly_budget REAL DEFAULT 0,
                created_at TEXT,
                updated_at TEXT
            )
        ''')
        
        # Ledger backup snapshots (created on ledger delete)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ledger_backups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ledger_id INTEGER,
                ledger_name TEXT,
                monthly_budget REAL,
                created_at TEXT,
                backup_json TEXT
            )
        ''')

        # Check if we need to migrate existing bills table
        cursor.execute("PRAGMA table_info(bills)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'bill_date' not in columns:
            # Migrate existing table
            self._migrate_bills_table(cursor)
        
        # Create enhanced bills table if it doesn't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_time TEXT,  -- Keep for backward compatibility
                image_name TEXT,
                merchant TEXT,
                category TEXT,
                amount REAL,
                raw_text TEXT,
                bill_date TEXT,
                created_at TEXT,
                updated_at TEXT,
                is_manual INTEGER DEFAULT 0,
                ledger_id INTEGER,
                include_in_budget INTEGER DEFAULT 1
            )
        ''')
        
        # Create category_rules table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS category_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT UNIQUE NOT NULL,
                category TEXT NOT NULL,
                priority INTEGER DEFAULT 1,
                is_weak INTEGER DEFAULT 0,
                created_at TEXT,
                updated_at TEXT,
                ledger_id INTEGER
            )
        ''')

        # Create categories table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                major TEXT NOT NULL,
                minor TEXT NOT NULL,
                created_at TEXT,
                updated_at TEXT,
                UNIQUE(major, minor)
            )
        ''')

        # Create recurring rules table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS recurring_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ledger_id INTEGER,
                amount REAL NOT NULL,
                keyword TEXT,
                category_id INTEGER,
                category TEXT,
                note TEXT,
                schedule_type TEXT NOT NULL,
                schedule_value INTEGER NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT,
                enabled INTEGER DEFAULT 1,
                include_in_budget INTEGER DEFAULT 1,
                created_at TEXT,
                updated_at TEXT
            )
        ''')

        cursor.execute("PRAGMA table_info(recurring_rules)")
        recurring_cols = [c[1] for c in cursor.fetchall()]
        if 'ledger_id' not in recurring_cols:
            try:
                cursor.execute('ALTER TABLE recurring_rules ADD COLUMN ledger_id INTEGER')
            except Exception:
                pass
        if 'keyword' not in recurring_cols:
            try:
                cursor.execute('ALTER TABLE recurring_rules ADD COLUMN keyword TEXT')
            except Exception:
                pass
        if 'include_in_budget' not in recurring_cols:
            try:
                cursor.execute('ALTER TABLE recurring_rules ADD COLUMN include_in_budget INTEGER DEFAULT 1')
            except Exception:
                pass

        # Track generated recurring bills to avoid duplicates
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS recurring_rule_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id INTEGER NOT NULL,
                ledger_id INTEGER,
                bill_date TEXT NOT NULL,
                bill_id INTEGER,
                created_at TEXT,
                UNIQUE(rule_id, bill_date)
            )
        ''')

        # Add ledger_id columns if missing
        for table in ['bills', 'category_rules', 'categories']:
            cursor.execute(f"PRAGMA table_info({table})")
            cols = [c[1] for c in cursor.fetchall()]
            if 'ledger_id' not in cols:
                try:
                    cursor.execute(f'ALTER TABLE {table} ADD COLUMN ledger_id INTEGER')
                except Exception:
                    pass

        cursor.execute("PRAGMA table_info(bills)")
        bill_cols = [c[1] for c in cursor.fetchall()]
        if 'include_in_budget' not in bill_cols:
            try:
                cursor.execute('ALTER TABLE bills ADD COLUMN include_in_budget INTEGER DEFAULT 1')
            except Exception:
                pass

        # Ensure default ledger exists before running migrations that need it
        cursor.execute('SELECT COUNT(*) FROM ledgers')
        if cursor.fetchone()[0] == 0:
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute('INSERT INTO ledgers (name, monthly_budget, created_at, updated_at) VALUES (?, ?, ?, ?)',
                           ('默认账本', 0, now, now))
            conn.commit()

        # Rebuild tables to add unique constraints with ledger_id (if not yet migrated)
        cursor.execute('PRAGMA user_version')
        user_version = cursor.fetchone()[0] or 0
        if user_version < 1:
            self._migrate_categories_with_ledger(cursor)
            self._migrate_category_rules_with_ledger(cursor)
            self._migrate_bills_with_ledger(cursor)
            cursor.execute('PRAGMA user_version = 1')
            user_version = 1
        if user_version < 2:
            self._migrate_categories_with_ids(cursor)
            cursor.execute('PRAGMA user_version = 2')
        
        # Create indexes for performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bills_bill_date ON bills(bill_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bills_category ON bills(category)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_bills_created_at ON bills(created_at)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_category_rules_keyword ON category_rules(keyword)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_categories_major_minor ON categories(major, minor)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_recurring_rules_ledger ON recurring_rules(ledger_id)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_recurring_rule_runs_rule_date ON recurring_rule_runs(rule_id, bill_date)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ledger_backups_ledger ON ledger_backups(ledger_id)')
        
        # Initialize category rules from config if table is empty
        cursor.execute('SELECT COUNT(*) FROM category_rules')
        if cursor.fetchone()[0] == 0:
            self._init_category_rules_from_config(cursor)
        self._ensure_config_rules_global(cursor)

        # Initialize categories from existing rules or config if table is empty
        cursor.execute('SELECT COUNT(*) FROM categories')
        if cursor.fetchone()[0] == 0:
            self._init_categories_from_rules_and_config(cursor)
        else:
            self._sync_categories_from_rules(cursor)

        # Cleanup legacy NULL-ledger duplicates once on startup
        self._cleanup_categories(cursor)
        
        conn.commit()
        conn.close()

    def _migrate_categories_with_ledger(self, cursor):
        cursor.execute("PRAGMA table_info(categories)")
        cols = cursor.fetchall()
        has_ledger = any(c[1] == 'ledger_id' for c in cols)
        cursor.execute("PRAGMA index_list(categories)")
        idxs = cursor.fetchall()
        has_ledger_unique = any('ledger' in (idx[1] or '') for idx in idxs)
        if has_ledger and has_ledger_unique:
            return

        cursor.execute("SELECT id, major, minor, created_at, updated_at, ledger_id FROM categories")
        rows = cursor.fetchall()
        cursor.execute("ALTER TABLE categories RENAME TO categories_old")
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                major TEXT NOT NULL,
                minor TEXT NOT NULL,
                created_at TEXT,
                updated_at TEXT,
                ledger_id INTEGER,
                UNIQUE(major, minor, ledger_id)
            )
        ''')
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row in rows:
            cursor.execute('''
                INSERT OR IGNORE INTO categories (id, major, minor, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (row[0], row[1], row[2], row[3] or now, row[4] or now, row[5] if len(row) > 5 else None))
        cursor.execute("DROP TABLE IF EXISTS categories_old")

    def _migrate_category_rules_with_ledger(self, cursor):
        cursor.execute("PRAGMA table_info(category_rules)")
        rows = cursor.fetchall()
        has_ledger = any(r[1] == 'ledger_id' for r in rows)
        cursor.execute("PRAGMA index_list(category_rules)")
        idxs = cursor.fetchall()
        has_ledger_unique = any('ledger' in (idx[1] or '') for idx in idxs)
        if has_ledger and has_ledger_unique:
            return
        cursor.execute("SELECT id, keyword, category, priority, is_weak, created_at, updated_at, ledger_id FROM category_rules")
        old_rows = cursor.fetchall()
        cursor.execute("ALTER TABLE category_rules RENAME TO category_rules_old")
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS category_rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT NOT NULL,
                category TEXT NOT NULL,
                priority INTEGER DEFAULT 1,
                is_weak INTEGER DEFAULT 0,
                created_at TEXT,
                updated_at TEXT,
                ledger_id INTEGER,
                UNIQUE(keyword, ledger_id)
            )
        ''')
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for r in old_rows:
            cursor.execute('''
                INSERT OR IGNORE INTO category_rules (id, keyword, category, priority, is_weak, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (r[0], r[1], r[2], r[3], r[4], r[5] or now, r[6] or now, r[7] if len(r) > 7 else None))
        cursor.execute("DROP TABLE IF EXISTS category_rules_old")

    def _migrate_bills_with_ledger(self, cursor):
        cursor.execute("PRAGMA table_info(bills)")
        cols = cursor.fetchall()
        has_ledger = any(c[1] == 'ledger_id' for c in cols)
        if not has_ledger:
            return
        # assign default ledger_id to NULL entries
        default_ledger = self.get_default_ledger_id()
        cursor.execute('UPDATE bills SET ledger_id = ? WHERE ledger_id IS NULL', (default_ledger,))

    def _migrate_categories_with_ids(self, cursor):
        # 0) 确认 bills/category_rules 不是 VIEW（否则 ALTER TABLE 会失败）
        for table in ("bills", "category_rules"):
            cursor.execute(
                "SELECT type, sql FROM sqlite_master WHERE name=? COLLATE NOCASE",
                (table,),
            )
            row = cursor.fetchone()
            if not row:
                raise RuntimeError(f"Table not found: {table}")
            if row[0] != "table":
                raise RuntimeError(f"{table} is not a table (type={row[0]}), cannot ALTER. sql={row[1]}")

        # 1) add category_id columns（不要吞异常）
        for table in ("bills", "category_rules"):
            cursor.execute(f"PRAGMA table_info({table})")
            cols = [c[1] for c in cursor.fetchall()]
            if "category_id" not in cols:
                cursor.execute(f"ALTER TABLE {table} ADD COLUMN category_id INTEGER")
                # 立刻验证
                cursor.execute(f"PRAGMA table_info({table})")
                cols2 = [c[1] for c in cursor.fetchall()]
                if "category_id" not in cols2:
                    raise RuntimeError(f"Failed to add category_id to {table}")

        # 2) build mapping: (major, minor, ledger_id)->category_id
        cursor.execute("SELECT id, major, minor, ledger_id FROM categories")
        categories = cursor.fetchall()

        cat_map = {}
        for cid, maj, mino, lid in categories:
            cat_map[(maj or "", mino or "", lid)] = cid
            # 如果你希望“ledger 维度匹配不到时可以回退到全局分类(ledger_id NULL)”
            # 可以额外放一个 fallback key
            if lid is None:
                cat_map[(maj or "", mino or "", None)] = cid

        def cat_id_by_name(name, ledger_id):
            name = (name or "").strip()
            major, minor = (name.split("/", 1) + [""])[:2]
            major = major.strip()
            minor = minor.strip()
            # 先找同 ledger 的分类
            cid = cat_map.get((major, minor, ledger_id))
            if cid:
                return cid
            # 再找全局分类（ledger_id NULL）
            return cat_map.get((major, minor, None))

        # 3) bills backfill（用 0 而不是 ''）
        cursor.execute(
            "SELECT id, category, ledger_id FROM bills WHERE category_id IS NULL OR category_id = 0"
        )
        for bid, cat, lid in cursor.fetchall():
            cid = cat_id_by_name(cat, lid)
            if cid:
                cursor.execute("UPDATE bills SET category_id=? WHERE id=?", (cid, bid))

        # 4) rules backfill
        cursor.execute(
            "SELECT id, category, ledger_id FROM category_rules WHERE category_id IS NULL OR category_id = 0"
        )
        for rid, cat, lid in cursor.fetchall():
            cid = cat_id_by_name(cat, lid)
            if cid:
                cursor.execute("UPDATE category_rules SET category_id=? WHERE id=?", (cid, rid))


    def get_default_ledger_id(self) -> int:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM ledgers ORDER BY id LIMIT 1')
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else 0

    def get_ledger_count(self) -> int:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM ledgers')
        count = cursor.fetchone()[0] or 0
        conn.close()
        return count

    def _table_columns(self, cursor, table: str) -> List[str]:
        cursor.execute(f"PRAGMA table_info({table})")
        return [row[1] for row in cursor.fetchall()]

    def _fetch_rows(self, cursor, table: str, where_clause: str = "", params: Optional[List[Any]] = None) -> List[Dict[str, Any]]:
        query = f"SELECT * FROM {table}"
        if where_clause:
            query += f" WHERE {where_clause}"
        cursor.execute(query, params or [])
        cols = [desc[0] for desc in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]

    def _insert_rows(self, cursor, table: str, rows: List[Dict[str, Any]]):
        if not rows:
            return
        available_cols = self._table_columns(cursor, table)
        row_keys = set()
        for row in rows:
            row_keys.update(row.keys())
        use_cols = [col for col in available_cols if col in row_keys]
        if not use_cols:
            return
        placeholders = ",".join(["?"] * len(use_cols))
        sql = f"INSERT INTO {table} ({', '.join(use_cols)}) VALUES ({placeholders})"
        for row in rows:
            cursor.execute(sql, [row.get(col) for col in use_cols])

    def _build_ledger_backup(self, cursor, ledger_id: int) -> Optional[Dict[str, Any]]:
        cursor.execute('SELECT id, name, monthly_budget, created_at, updated_at FROM ledgers WHERE id=?', (ledger_id,))
        ledger_row = cursor.fetchone()
        if not ledger_row:
            return None
        ledger = {
            "id": ledger_row[0],
            "name": ledger_row[1],
            "monthly_budget": ledger_row[2] or 0,
            "created_at": ledger_row[3],
            "updated_at": ledger_row[4],
        }
        return {
            "version": 1,
            "ledger": ledger,
            "categories": self._fetch_rows(cursor, "categories", "ledger_id = ?", [ledger_id]),
            "category_rules": self._fetch_rows(cursor, "category_rules", "ledger_id = ?", [ledger_id]),
            "bills": self._fetch_rows(cursor, "bills", "ledger_id = ?", [ledger_id]),
            "recurring_rules": self._fetch_rows(cursor, "recurring_rules", "ledger_id = ?", [ledger_id]),
            "recurring_rule_runs": self._fetch_rows(cursor, "recurring_rule_runs", "ledger_id = ?", [ledger_id]),
        }

    def create_ledger_backup(self, ledger_id: int, cursor: Optional[sqlite3.Cursor] = None) -> Optional[int]:
        owns_conn = cursor is None
        conn = None
        if owns_conn:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
        backup = self._build_ledger_backup(cursor, ledger_id)
        if not backup:
            if owns_conn and conn is not None:
                conn.close()
            return None
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        backup_json = json.dumps(backup, ensure_ascii=False)
        cursor.execute(
            '''
            INSERT INTO ledger_backups (ledger_id, ledger_name, monthly_budget, created_at, backup_json)
            VALUES (?, ?, ?, ?, ?)
            ''',
            (backup["ledger"]["id"], backup["ledger"]["name"], backup["ledger"]["monthly_budget"], now, backup_json),
        )
        backup_id = cursor.lastrowid
        if owns_conn and conn is not None:
            conn.commit()
            conn.close()
        return backup_id

    def list_ledger_backups(self) -> List[Dict[str, Any]]:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute(
            '''
            SELECT id, ledger_id, ledger_name, monthly_budget, created_at
            FROM ledger_backups
            ORDER BY id DESC
            '''
        )
        rows = cursor.fetchall()
        conn.close()
        return [
            {
                "id": row[0],
                "ledger_id": row[1],
                "ledger_name": row[2],
                "monthly_budget": row[3] or 0,
                "created_at": row[4],
            }
            for row in rows
        ]

    def delete_ledger_backup(self, backup_id: int) -> bool:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM ledger_backups WHERE id=?', (backup_id,))
        deleted = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return deleted

    def restore_ledger_backup(self, backup_id: int) -> Dict[str, Any]:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute(
            '''
            SELECT id, ledger_id, ledger_name, monthly_budget, created_at, backup_json
            FROM ledger_backups
            WHERE id = ?
            ''',
            (backup_id,),
        )
        row = cursor.fetchone()
        if not row:
            conn.close()
            return {"success": False, "error": "backup_not_found"}

        try:
            backup = json.loads(row[5] or "{}")
        except json.JSONDecodeError:
            conn.close()
            return {"success": False, "error": "backup_invalid"}

        ledger = backup.get("ledger") or {}
        ledger_id = ledger.get("id") or row[1]
        ledger_name = ledger.get("name") or row[2] or f"Ledger {ledger_id}"
        monthly_budget = ledger.get("monthly_budget") or row[3] or 0
        created_at = ledger.get("created_at") or row[4] or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated_at = ledger.get("updated_at") or created_at

        cursor.execute('SELECT COUNT(*) FROM ledgers WHERE id=?', (ledger_id,))
        if cursor.fetchone()[0] > 0:
            conn.close()
            return {"success": False, "error": "ledger_exists"}

        cursor.execute('SELECT COUNT(*) FROM ledgers WHERE name=?', (ledger_name,))
        if cursor.fetchone()[0] > 0:
            suffix = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            ledger_name = f"{ledger_name} (恢复{suffix})"

        try:
            conn.execute("BEGIN")
            # Clean any stale rows for this ledger id (defensive)
            for table in ("bills", "category_rules", "categories", "recurring_rules", "recurring_rule_runs"):
                cursor.execute(f"DELETE FROM {table} WHERE ledger_id = ?", (ledger_id,))

            cursor.execute(
                '''
                INSERT INTO ledgers (id, name, monthly_budget, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ''',
                (ledger_id, ledger_name, monthly_budget, created_at, updated_at),
            )

            def apply_ledger_id(rows: List[Dict[str, Any]]):
                for r in rows:
                    if "ledger_id" in r:
                        r["ledger_id"] = ledger_id

            categories = backup.get("categories", [])
            rules = backup.get("category_rules", [])
            bills = backup.get("bills", [])
            recurring_rules = backup.get("recurring_rules", [])
            recurring_runs = backup.get("recurring_rule_runs", [])

            apply_ledger_id(categories)
            apply_ledger_id(rules)
            apply_ledger_id(bills)
            apply_ledger_id(recurring_rules)
            apply_ledger_id(recurring_runs)

            self._insert_rows(cursor, "categories", categories)
            self._insert_rows(cursor, "category_rules", rules)
            self._insert_rows(cursor, "bills", bills)
            self._insert_rows(cursor, "recurring_rules", recurring_rules)
            self._insert_rows(cursor, "recurring_rule_runs", recurring_runs)

            conn.commit()
        except Exception:
            conn.rollback()
            conn.close()
            return {"success": False, "error": "restore_failed"}

        conn.close()
        return {"success": True, "ledger_id": ledger_id, "ledger_name": ledger_name}

    def list_ledgers(self) -> List[Dict[str, Any]]:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, monthly_budget, created_at, updated_at FROM ledgers ORDER BY id')
        rows = cursor.fetchall()
        conn.close()
        return [
            {
                "id": row[0],
                "name": row[1],
                "monthly_budget": row[2] or 0,
                "created_at": row[3],
                "updated_at": row[4]
            }
            for row in rows
        ]

    def save_ledger(self, ledger: Dict[str, Any]) -> int:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ledger_id = ledger.get("id")
        name = ledger.get("name")
        budget = ledger.get("monthly_budget") or 0
        if ledger_id:
            cursor.execute('UPDATE ledgers SET name=?, monthly_budget=?, updated_at=? WHERE id=?',
                           (name, budget, now, ledger_id))
        else:
            cursor.execute('INSERT INTO ledgers (name, monthly_budget, created_at, updated_at) VALUES (?, ?, ?, ?)',
                           (name, budget, now, now))
            ledger_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return ledger_id

    def delete_ledger(self, ledger_id: int) -> bool:
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM ledgers')
        total = cursor.fetchone()[0] or 0
        if total <= 1:
            conn.close()
            return False

        backup_id = self.create_ledger_backup(ledger_id, cursor=cursor)
        if backup_id is None:
            conn.close()
            return False

        cursor.execute('DELETE FROM recurring_rule_runs WHERE ledger_id=?', (ledger_id,))
        cursor.execute('DELETE FROM recurring_rules WHERE ledger_id=?', (ledger_id,))
        cursor.execute('DELETE FROM bills WHERE ledger_id=?', (ledger_id,))
        cursor.execute('DELETE FROM category_rules WHERE ledger_id=?', (ledger_id,))
        cursor.execute('DELETE FROM categories WHERE ledger_id=?', (ledger_id,))
        cursor.execute('DELETE FROM ledgers WHERE id=?', (ledger_id,))
        deleted = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return deleted
        print(f"📘 [Enhanced DB] 数据库连接就绪: {self.db_name}")
    
    def _migrate_bills_table(self, cursor):
        """Migrate existing bills table to new schema"""
        print("🔄 [DB Migration] 正在升级数据库架构...")
        
        # Add new columns to existing table
        try:
            cursor.execute('ALTER TABLE bills ADD COLUMN bill_date TEXT')
            cursor.execute('ALTER TABLE bills ADD COLUMN created_at TEXT')
            cursor.execute('ALTER TABLE bills ADD COLUMN updated_at TEXT')
            cursor.execute('ALTER TABLE bills ADD COLUMN is_manual INTEGER DEFAULT 0')
            
            # Update existing records with default values
            cursor.execute('''
                UPDATE bills 
                SET bill_date = date(record_time),
                    created_at = record_time,
                    updated_at = record_time
                WHERE bill_date IS NULL
            ''')
            
            print("✅ [DB Migration] 数据库架构升级完成")
        except sqlite3.OperationalError as e:
            if "duplicate column name" not in str(e):
                raise e
    
    def _init_category_rules_from_config(self, cursor):
        """Initialize category rules from config file"""
        print("🔄 [DB Init] 正在从配置文件初始化分类规则...")
        
        for keyword, category in config.CATEGORY_RULES.items():
            priority = 2
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            cursor.execute('''
                INSERT INTO category_rules (keyword, category, priority, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?, NULL)
            ''', (keyword, category, priority, now, now))
        
        print("✅ [DB Init] 分类规则初始化完成")

    def _ensure_config_rules_global(self, cursor):
        """Ensure config-based rules are stored as global (ledger_id NULL)."""
        if not config.CATEGORY_RULES:
            return
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for keyword, category in config.CATEGORY_RULES.items():
            cursor.execute(
                "SELECT id FROM category_rules WHERE keyword=? AND category=? AND ledger_id IS NULL",
                (keyword, category),
            )
            if cursor.fetchone():
                continue
            cursor.execute(
                "SELECT id FROM category_rules WHERE keyword=? AND category=? ORDER BY id LIMIT 1",
                (keyword, category),
            )
            row = cursor.fetchone()
            if row:
                cursor.execute("UPDATE category_rules SET ledger_id=NULL WHERE id=?", (row[0],))
                continue
            cursor.execute(
                '''
                INSERT INTO category_rules (keyword, category, priority, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?, NULL)
                ''',
                (keyword, category, 2, now, now),
            )

    def _format_category_name(self, major: str, minor: str) -> str:
        major = (major or "").strip()
        minor = (minor or "").strip()
        return f"{major}/{minor}" if minor else major

    def _split_category_name(self, category_name: str) -> CategoryGroup:
        if not category_name:
            return CategoryGroup(major="", minor="")
        parts = category_name.split('/', 1)
        if len(parts) == 2:
            return CategoryGroup(major=parts[0].strip(), minor=parts[1].strip())
        return CategoryGroup(major=category_name.strip(), minor="")

    def _normalize_schedule_values(self, value: Any) -> List[int]:
        if value is None:
            return []
        if isinstance(value, list):
            values = value
        elif isinstance(value, str):
            text = value.strip()
            values = [int(v) for v in text.split(',') if v.strip()] if text else []
        else:
            values = [value]
        normalized = []
        for v in values:
            try:
                num = int(v)
                normalized.append(num)
            except (TypeError, ValueError):
                continue
        return sorted(set(normalized))

    def _resolve_category(self, cursor, category_name: str, category_id: Optional[int], ledger_id: Optional[int]):
        """Return (category_id, category_full_name) using ID when provided; otherwise try to find by name."""
        if category_id:
            cursor.execute("SELECT major, minor FROM categories WHERE id=?", (category_id,))
            row = cursor.fetchone()
            if row:
                return category_id, self._format_category_name(row[0], row[1])
        if not category_name:
            return None, ""
        group = self._split_category_name(category_name)
        # prefer ledger-specific match, fallback to global
        params = [group.major, group.minor]
        query = "SELECT id, major, minor FROM categories WHERE major=? AND minor=?"
        if ledger_id is None:
            query += " AND ledger_id IS NULL"
        else:
            query += " AND (ledger_id IS NULL OR ledger_id = ?)"
            params.append(ledger_id)
        cursor.execute(query, params)
        row = cursor.fetchone()
        if row:
            return row[0], self._format_category_name(row[1], row[2])
        return None, category_name

    def _init_categories_from_rules_and_config(self, cursor):
        """Initialize categories from existing rules and config"""
        print("🔄 [DB Init] 正在初始化分类目录...")
        categories = set()
        cursor.execute('SELECT DISTINCT category FROM category_rules')
        for row in cursor.fetchall():
            if row[0]:
                categories.add(row[0])

        for category in set(config.CATEGORY_RULES.values()):
            if category:
                categories.add(category)

        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for category_name in categories:
            group = self._split_category_name(category_name)
            if not group.major:
                continue
            cursor.execute('''
                INSERT OR IGNORE INTO categories (major, minor, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?)
            ''', (group.major, group.minor, now, now, None))

        print("✅ [DB Init] 分类目录初始化完成")

    def _sync_categories_from_rules(self, cursor):
        """Ensure categories table covers categories referenced by rules"""
        cursor.execute('SELECT DISTINCT category, ledger_id FROM category_rules')
        rows = cursor.fetchall()
        if not rows:
            return
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for category_name, lid in rows:
            if not category_name:
                continue
            group = self._split_category_name(category_name)
            if not group.major:
                continue
            cursor.execute('''
                INSERT OR IGNORE INTO categories (major, minor, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?)
            ''', (group.major, group.minor, now, now, lid))

    def _cleanup_categories(self, cursor):
        """Normalize ledger_id on categories and remove duplicates created before migration."""
        cursor.execute('UPDATE categories SET ledger_id = NULL WHERE ledger_id = 0')
        cursor.execute('SELECT id, major, minor, ledger_id FROM categories')
        rows = cursor.fetchall()
        if not rows:
            return

        rows_sorted = sorted(rows, key=lambda r: r[0])
        keepers = {}
        duplicates = []
        for cid, major, minor, lid in rows_sorted:
            normalized_ledger = None if lid in (0,) else lid
            key = ((major or "").strip(), (minor or "").strip(), normalized_ledger)
            if key in keepers:
                duplicates.append((cid, keepers[key]))
            else:
                keepers[key] = cid

        for dup_id, keep_id in duplicates:
            cursor.execute('UPDATE bills SET category_id=? WHERE category_id=?', (keep_id, dup_id))
            cursor.execute('UPDATE category_rules SET category_id=? WHERE category_id=?', (keep_id, dup_id))

        if duplicates:
            dup_ids = [d[0] for d in duplicates]
            placeholders = ','.join(['?'] * len(dup_ids))
            cursor.execute(f'DELETE FROM categories WHERE id IN ({placeholders})', dup_ids)

        cursor.execute("""
            SELECT DISTINCT category_id
            FROM category_rules
            WHERE ledger_id IS NULL AND category_id IS NOT NULL AND category_id != 0
        """)
        global_rule_category_ids = [row[0] for row in cursor.fetchall()]
        for cat_id in global_rule_category_ids:
            cursor.execute('SELECT id, major, minor, ledger_id FROM categories WHERE id=?', (cat_id,))
            row = cursor.fetchone()
            if not row:
                continue
            if row[3] is None:
                continue
            cursor.execute(
                'SELECT id FROM categories WHERE major=? AND minor=? AND ledger_id IS NULL',
                (row[1], row[2]),
            )
            global_row = cursor.fetchone()
            if global_row:
                global_id = global_row[0]
                cursor.execute(
                    'UPDATE category_rules SET category_id=? WHERE ledger_id IS NULL AND category_id=?',
                    (global_id, cat_id),
                )
            else:
                try:
                    cursor.execute('UPDATE categories SET ledger_id=NULL WHERE id=?', (cat_id,))
                except sqlite3.IntegrityError:
                    cursor.execute(
                        'SELECT id FROM categories WHERE major=? AND minor=? AND ledger_id IS NULL',
                        (row[1], row[2]),
                    )
                    fallback = cursor.fetchone()
                    if fallback:
                        cursor.execute(
                            'UPDATE category_rules SET category_id=? WHERE ledger_id IS NULL AND category_id=?',
                            (fallback[0], cat_id),
                        )

    # Bills CRUD operations
    def save_bill(self, bill: EnhancedBill) -> int:
        """Save a bill to database"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        if bill.ledger_id is None:
            bill.ledger_id = self.get_default_ledger_id()
        # resolve category by id or name
        bill.category_id, bill.category = self._resolve_category(cursor, bill.category, bill.category_id, bill.ledger_id)
        bill.updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        raw_text_json = json.dumps(bill.raw_text, ensure_ascii=False)
        
        if bill.id:
            # Update existing bill
            cursor.execute('''
                UPDATE bills 
                SET image_name=?, merchant=?, category=?, category_id=?, amount=?, raw_text=?,
                    bill_date=?, updated_at=?, is_manual=?, ledger_id=?, include_in_budget=?
                WHERE id=?
            ''', (bill.filename, bill.merchant, bill.category, bill.category_id, bill.amount, 
                  raw_text_json, bill.bill_date, bill.updated_at, 
                  int(bill.is_manual), bill.ledger_id, int(bool(bill.include_in_budget)), bill.id))
            bill_id = bill.id
        else:
            # Insert new bill
            cursor.execute('''
                INSERT INTO bills (record_time, image_name, merchant, category, category_id, amount, 
                                 raw_text, bill_date, created_at, updated_at, is_manual, ledger_id, include_in_budget)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (bill.created_at, bill.filename, bill.merchant, bill.category, bill.category_id,
                  bill.amount, raw_text_json, bill.bill_date, bill.created_at, 
                  bill.updated_at, int(bill.is_manual), bill.ledger_id, int(bool(bill.include_in_budget))))
            bill_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        return bill_id
    
    def get_bill(self, bill_id: int) -> Optional[EnhancedBill]:
        """Get a bill by ID"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM bills WHERE id = ?', (bill_id,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return self._row_to_bill(row)
        return None
    
    def _append_category_filter(self, query: str, params: List[Any], categories: Optional[List[str]], col_expr: str = "category") -> str:
        if categories is None:
            return query
        if not categories:
            return f"{query} AND 0"
        placeholders = ','.join(['?'] * len(categories))
        query += f' AND {col_expr} IN ({placeholders})'
        params.extend(categories)
        return query

    def _fetch_category_names_by_major_minor(self, cursor, major: str = None, minor: str = None, ledger_id: Optional[int] = None) -> Optional[List[str]]:
        if not major and not minor:
            return None
        query = 'SELECT major, minor FROM categories WHERE 1=1'
        params: List[Any] = []
        if major:
            query += ' AND major = ?'
            params.append(major)
        if minor:
            query += ' AND minor = ?'
            params.append(minor)
        if ledger_id is not None:
            query += ' AND (ledger_id IS NULL OR ledger_id = ?)'
            params.append(ledger_id)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return [self._format_category_name(row[0], row[1]) for row in rows]

    def get_bills(self, limit: int = 100, offset: int = 0,
                  start_date: str = None, end_date: str = None,
                  category: str = None, keyword: str = None,
                  major: str = None, minor: str = None,
                  ledger_id: Optional[int] = None,
                  sort_by: Optional[str] = None,
                  sort_order: Optional[str] = None) -> List[EnhancedBill]:
        """Get bills with filtering and pagination"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        query = '''
            SELECT b.*, c.major, c.minor
            FROM bills b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE 1=1
        '''
        params = []
        
        if start_date:
            query += ' AND b.bill_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND b.bill_date <= ?'
            params.append(end_date)
        
        categories_filter = None
        if category:
            categories_filter = [category]
        elif major or minor:
            categories_filter = self._fetch_category_names_by_major_minor(cursor, major, minor, ledger_id)
        query = self._append_category_filter(query, params, categories_filter, "COALESCE(c.major || '/' || c.minor, b.category)")

        if keyword:
            keyword_like = f'%{keyword}%'
            query += ' AND (b.merchant LIKE ? OR b.raw_text LIKE ?)'
            params.extend([keyword_like, keyword_like])

        if ledger_id is not None:
            query += ' AND b.ledger_id = ?'
            params.append(ledger_id)

        sort_map = {
            'bill_date': 'b.bill_date',
            'merchant': 'b.merchant',
            'category': "COALESCE(c.major || '/' || c.minor, b.category)",
            'amount': 'b.amount',
        }
        sort_field = sort_map.get((sort_by or '').strip())
        sort_dir = 'ASC' if (sort_order or '').lower() == 'asc' else 'DESC'
        if sort_field:
            query += f' ORDER BY {sort_field} {sort_dir}, b.created_at DESC'
        else:
            query += ' ORDER BY b.bill_date DESC, b.created_at DESC'

        query += ' LIMIT ? OFFSET ?'
        params.extend([limit, offset])
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        return [self._row_to_bill(row) for row in rows]

    def get_bills_count(self, start_date: str = None, end_date: str = None,
                        category: str = None, keyword: str = None,
                        major: str = None, minor: str = None,
                        ledger_id: Optional[int] = None) -> int:
        """Get total bill count with filtering"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        query = '''
            SELECT COUNT(*)
            FROM bills b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE 1=1
        '''
        params = []

        if start_date:
            query += ' AND b.bill_date >= ?'
            params.append(start_date)

        if end_date:
            query += ' AND b.bill_date <= ?'
            params.append(end_date)

        categories_filter = None
        if category:
            categories_filter = [category]
        elif major or minor:
            categories_filter = self._fetch_category_names_by_major_minor(cursor, major, minor, ledger_id)
        query = self._append_category_filter(query, params, categories_filter, "COALESCE(c.major || '/' || c.minor, b.category)")

        if keyword:
            keyword_like = f'%{keyword}%'
            query += ' AND (b.merchant LIKE ? OR b.raw_text LIKE ?)'
            params.extend([keyword_like, keyword_like])

        if ledger_id is not None:
            query += ' AND b.ledger_id = ?'
            params.append(ledger_id)

        cursor.execute(query, params)
        count = cursor.fetchone()[0] or 0
        conn.close()
        return count
    
    def delete_bill(self, bill_id: int) -> bool:
        """Delete a bill by ID"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM bills WHERE id = ?', (bill_id,))
        deleted = cursor.rowcount > 0
        
        conn.commit()
        conn.close()
        return deleted
    
    def update_bill_budget_status(self, bill_id: int, include_in_budget: bool) -> bool:
        """Update a bill's budget inclusion status"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute(
            'UPDATE bills SET include_in_budget = ? WHERE id = ?', 
            (include_in_budget, bill_id)
        )
        updated = cursor.rowcount > 0
        
        conn.commit()
        conn.close()
        return updated
    
    def _row_to_bill(self, row) -> EnhancedBill:
        """Convert database row to EnhancedBill object"""
        raw_text = json.loads(row[6]) if row[6] else []
        cat_id = row[12] if len(row) > 12 else None
        include_in_budget = bool(row[13]) if len(row) > 13 and row[13] is not None else True
        cat_major = row[14] if len(row) > 14 else None
        cat_minor = row[15] if len(row) > 15 else None
        category_name = self._format_category_name(cat_major, cat_minor) if cat_major is not None else (row[4] or "")
        
        return EnhancedBill(
            id=row[0],
            ledger_id=row[11] if len(row) > 11 else None,
            filename=row[2] or "",
            merchant=row[3] or "",
            amount=row[5] or 0.0,
            category_id=cat_id,
            category=category_name,
            bill_date=row[7] or datetime.date.today().strftime("%Y-%m-%d"),
            created_at=row[8] or row[1],  # Fallback to record_time
            updated_at=row[9] or row[1],  # Fallback to record_time
            raw_text=raw_text,
            is_manual=bool(row[10]) if len(row) > 10 else False,
            include_in_budget=include_in_budget
        )
    
    # Category Rules CRUD operations
    def get_category_rules(self, ledger_id: Optional[int] = None) -> List[CategoryRule]:
        """Get all category rules (global + specific ledger when provided)"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        query = '''
            SELECT r.id, r.keyword, r.category, r.category_id, r.priority,
                   r.created_at, r.updated_at, r.ledger_id, c.major, c.minor
            FROM category_rules r
            LEFT JOIN categories c ON r.category_id = c.id
            WHERE 1=1
        '''
        params: List[Any] = []
        if ledger_id is not None:
            query += ' AND (r.ledger_id IS NULL OR r.ledger_id = ?)'
            params.append(ledger_id)

        query += ' ORDER BY r.category, r.priority DESC, r.keyword'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        return [self._row_to_category_rule(row) for row in rows]
    
    def save_category_rule(self, rule: CategoryRule) -> int:
        """Save a category rule"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        rule.category_id, rule.category = self._resolve_category(cursor, rule.category, rule.category_id, rule.ledger_id)
        rule.updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if rule.id:
            # Update existing rule
            cursor.execute('''
                UPDATE category_rules 
                SET keyword=?, category=?, category_id=?, priority=?, updated_at=?, ledger_id=?
                WHERE id=?
            ''', (rule.keyword, rule.category, rule.category_id, rule.priority,
                  rule.updated_at, rule.ledger_id, rule.id))
            rule_id = rule.id
        else:
            # Insert new rule
            cursor.execute('''
                INSERT INTO category_rules (keyword, category, category_id, priority, created_at, updated_at, ledger_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (rule.keyword, rule.category, rule.category_id, rule.priority,
                  rule.created_at, rule.updated_at, rule.ledger_id))
            rule_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        return rule_id
    
    def delete_category_rule(self, rule_id: int) -> bool:
        """Delete a category rule"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM category_rules WHERE id = ?', (rule_id,))
        deleted = cursor.rowcount > 0
        
        conn.commit()
        conn.close()
        return deleted

    # Recurring Rules CRUD operations
    def get_recurring_rules(self, ledger_id: Optional[int] = None) -> List[RecurringRule]:
        """Get recurring rules for a ledger"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        if ledger_id is None:
            ledger_id = self.get_default_ledger_id()

        query = '''
            SELECT r.id, r.ledger_id, r.amount, r.keyword, r.category_id, r.category, r.note,
                   r.schedule_type, r.schedule_value, r.start_date, r.end_date,
                   r.enabled, r.include_in_budget, r.created_at, r.updated_at, c.major, c.minor
            FROM recurring_rules r
            LEFT JOIN categories c ON r.category_id = c.id
            WHERE r.ledger_id = ?
            ORDER BY r.enabled DESC, r.start_date, r.id DESC
        '''
        cursor.execute(query, (ledger_id,))
        rows = cursor.fetchall()
        conn.close()
        return [self._row_to_recurring_rule(row) for row in rows]

    def get_recurring_rule(self, rule_id: int) -> Optional[RecurringRule]:
        """Get a recurring rule by id"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        query = '''
            SELECT r.id, r.ledger_id, r.amount, r.keyword, r.category_id, r.category, r.note,
                   r.schedule_type, r.schedule_value, r.start_date, r.end_date,
                   r.enabled, r.include_in_budget, r.created_at, r.updated_at, c.major, c.minor
            FROM recurring_rules r
            LEFT JOIN categories c ON r.category_id = c.id
            WHERE r.id = ?
        '''
        cursor.execute(query, (rule_id,))
        row = cursor.fetchone()
        conn.close()
        return self._row_to_recurring_rule(row) if row else None

    def save_recurring_rule(self, rule: RecurringRule) -> int:
        """Save a recurring rule"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        if rule.ledger_id is None:
            rule.ledger_id = self.get_default_ledger_id()

        rule.amount = float(rule.amount or 0)
        rule.schedule_value = self._normalize_schedule_values(rule.schedule_value)
        if not rule.schedule_value:
            rule.schedule_value = [1]
        rule.enabled = bool(rule.enabled)
        rule.include_in_budget = rule.include_in_budget if rule.include_in_budget is not None else True
        rule.category_id, rule.category = self._resolve_category(
            cursor, rule.category, rule.category_id, rule.ledger_id
        )
        rule.updated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        schedule_value_str = ",".join(str(v) for v in rule.schedule_value)

        if rule.id:
            cursor.execute('''
                UPDATE recurring_rules
                SET amount=?, keyword=?, category_id=?, category=?, note=?, schedule_type=?, schedule_value=?,
                    start_date=?, end_date=?, enabled=?, include_in_budget=?, updated_at=?, ledger_id=?
                WHERE id=?
            ''', (
                rule.amount,
                rule.keyword,
                rule.category_id,
                rule.category,
                rule.note,
                rule.schedule_type,
                schedule_value_str,
                rule.start_date,
                rule.end_date,
                1 if rule.enabled else 0,
                1 if rule.include_in_budget else 0,
                rule.updated_at,
                rule.ledger_id,
                rule.id,
            ))
            rule_id = rule.id
        else:
            cursor.execute('''
                INSERT INTO recurring_rules (
                    ledger_id, amount, keyword, category_id, category, note, schedule_type, schedule_value,
                    start_date, end_date, enabled, include_in_budget, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                rule.ledger_id,
                rule.amount,
                rule.keyword,
                rule.category_id,
                rule.category,
                rule.note,
                rule.schedule_type,
                schedule_value_str,
                rule.start_date,
                rule.end_date,
                1 if rule.enabled else 0,
                1 if rule.include_in_budget else 0,
                rule.created_at,
                rule.updated_at,
            ))
            rule_id = cursor.lastrowid

        conn.commit()
        conn.close()
        return rule_id

    def delete_recurring_rule(self, rule_id: int) -> bool:
        """Delete a recurring rule"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM recurring_rules WHERE id = ?', (rule_id,))
        deleted = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return deleted

    def generate_recurring_bills(self, ledger_id: Optional[int] = None, upto_date: Optional[str] = None) -> int:
        """Generate bills from recurring rules up to a date (inclusive)."""
        if ledger_id is None:
            ledger_id = self.get_default_ledger_id()

        if upto_date:
            try:
                end_date = datetime.datetime.strptime(upto_date, "%Y-%m-%d").date()
            except ValueError:
                end_date = datetime.date.today()
        else:
            end_date = datetime.date.today()

        rules = [r for r in self.get_recurring_rules(ledger_id) if r.enabled]
        if not rules:
            return 0

        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        created = 0
        now_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for rule in rules:
            schedule_values = self._normalize_schedule_values(rule.schedule_value)
            limit = 7 if rule.schedule_type == "weekly" else 31
            schedule_values = [v for v in schedule_values if 1 <= v <= limit]
            if not schedule_values:
                continue

            try:
                start_date = datetime.datetime.strptime(rule.start_date, "%Y-%m-%d").date()
            except (TypeError, ValueError):
                start_date = end_date

            rule_end = end_date
            if rule.end_date:
                try:
                    rule_end = datetime.datetime.strptime(rule.end_date, "%Y-%m-%d").date()
                except ValueError:
                    rule_end = end_date

            if rule_end > end_date:
                rule_end = end_date
            if start_date > rule_end:
                continue

            current = start_date
            while current <= rule_end:
                match = False
                if rule.schedule_type == "weekly":
                    match = current.isoweekday() in schedule_values
                elif rule.schedule_type == "monthly":
                    match = current.day in schedule_values
                if match:
                    bill_date = current.strftime("%Y-%m-%d")
                    cursor.execute(
                        '''
                        INSERT OR IGNORE INTO recurring_rule_runs (rule_id, ledger_id, bill_date, created_at)
                        VALUES (?, ?, ?, ?)
                        ''',
                        (rule.id, ledger_id, bill_date, now_ts),
                    )
                    if cursor.rowcount > 0:
                        category_id, category_name = self._resolve_category(
                            cursor, rule.category, rule.category_id, ledger_id
                        )
                        merchant = rule.keyword or rule.note or "周期性账单"
                        raw_text_json = json.dumps([], ensure_ascii=False)
                        cursor.execute(
                            '''
                            INSERT INTO bills (record_time, image_name, merchant, category, category_id, amount,
                                              raw_text, bill_date, created_at, updated_at, is_manual, ledger_id,
                                              include_in_budget)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''',
                            (
                                now_ts,
                                "",
                                merchant,
                                category_name,
                                category_id,
                                float(rule.amount or 0),
                                raw_text_json,
                                bill_date,
                                now_ts,
                                now_ts,
                                0,
                                ledger_id,
                                1 if rule.include_in_budget else 0,
                            ),
                        )
                        bill_id = cursor.lastrowid
                        cursor.execute(
                            '''
                            UPDATE recurring_rule_runs
                            SET bill_id=?
                            WHERE rule_id=? AND bill_date=?
                            ''',
                            (bill_id, rule.id, bill_date),
                        )
                        created += 1
                current += datetime.timedelta(days=1)

        conn.commit()
        conn.close()
        return created
    
    def _row_to_category_rule(self, row) -> CategoryRule:
        """Convert database row to CategoryRule object"""
        cat_name = self._format_category_name(row[8], row[9]) if len(row) > 9 and row[8] is not None else row[2]
        return CategoryRule(
            id=row[0],
            keyword=row[1],
            category=cat_name,
            category_id=row[3] if len(row) > 3 else None,
            priority=row[4],
            is_weak=False,
            created_at=row[5],
            updated_at=row[6],
            ledger_id=row[7] if len(row) > 7 else None
        )

    def _row_to_recurring_rule(self, row) -> RecurringRule:
        if not row:
            return None
        cat_name = self._format_category_name(row[15], row[16]) if len(row) > 16 and row[15] is not None else row[5]
        return RecurringRule(
            id=row[0],
            ledger_id=row[1],
            amount=row[2] or 0.0,
            keyword=row[3] or "",
            category_id=row[4],
            category=cat_name,
            note=row[6] or "",
            schedule_type=row[7] or "weekly",
            schedule_value=self._normalize_schedule_values(row[8]) if len(row) > 8 else [1],
            start_date=row[9],
            end_date=row[10],
            enabled=bool(row[11]) if len(row) > 11 else True,
            include_in_budget=bool(row[12]) if len(row) > 12 and row[12] is not None else True,
            created_at=row[13] if len(row) > 13 else "",
            updated_at=row[14] if len(row) > 14 else "",
        )

    # Category Groups CRUD operations
    def get_category_groups(self, ledger_id: Optional[int] = None) -> List[CategoryGroup]:
        """Get all category groups (global + specific ledger)"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        params: List[Any] = []
        query = 'SELECT * FROM categories WHERE 1=1'
        if ledger_id is not None:
            query += ' AND (ledger_id IS NULL OR ledger_id = ?)'
            params.append(ledger_id)
        query += ' ORDER BY major, minor'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return [self._row_to_category_group(row) for row in rows]

    def get_category_group(self, category_id: int) -> Optional[CategoryGroup]:
        """Get a category group by ID"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM categories WHERE id = ?', (category_id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            return self._row_to_category_group(row)
        return None

    def get_category_group_by_name(self, category_name: str, ledger_id: Optional[int] = None) -> Optional[CategoryGroup]:
        """Get a category group by full name"""
        group = self._split_category_name(category_name)
        if not group.major:
            return None
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        params: List[Any] = [group.major, group.minor]
        query = 'SELECT * FROM categories WHERE major = ? AND minor = ?'
        if ledger_id is not None:
            query += ' AND (ledger_id IS NULL OR ledger_id = ?)'
            params.append(ledger_id)
        cursor.execute(query, params)
        row = cursor.fetchone()
        conn.close()
        if row:
            return self._row_to_category_group(row)
        return None

    def category_exists(self, category_name: str, ledger_id: Optional[int] = None) -> bool:
        """Check if a category exists"""
        return self.get_category_group_by_name(category_name, ledger_id) is not None

    def save_category_group(self, group: CategoryGroup) -> int:
        """Save a category group and propagate renames to rules"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if group.id:
            cursor.execute('SELECT major, minor, ledger_id FROM categories WHERE id = ?', (group.id,))
            row = cursor.fetchone()
            if not row:
                conn.close()
                raise ValueError("Category not found")

            old_name = self._format_category_name(row[0], row[1])
            new_name = self._format_category_name(group.major, group.minor)
            group.updated_at = now

            try:
                cursor.execute('''
                    UPDATE categories
                    SET major = ?, minor = ?, updated_at = ?, ledger_id = ?
                    WHERE id = ?
                ''', (group.major, group.minor, group.updated_at, group.ledger_id, group.id))
            except sqlite3.IntegrityError:
                conn.close()
                raise ValueError("Category already exists")

            if old_name != new_name:
                cursor.execute('''
                    UPDATE category_rules
                    SET category = ?
                    WHERE category = ? AND (ledger_id IS NULL OR ledger_id = ?)
                ''', (new_name, old_name, group.ledger_id))
                cursor.execute('''
                    UPDATE bills
                    SET category = ?
                    WHERE category = ? AND ledger_id = ?
                ''', (new_name, old_name, group.ledger_id))
            group_id = group.id
        else:
            group.created_at = now
            group.updated_at = now
            try:
                cursor.execute('''
                    INSERT INTO categories (major, minor, created_at, updated_at, ledger_id)
                    VALUES (?, ?, ?, ?, ?)
                ''', (group.major, group.minor, group.created_at, group.updated_at, group.ledger_id))
            except sqlite3.IntegrityError:
                conn.close()
                raise ValueError("Category already exists")
            group_id = cursor.lastrowid

        conn.commit()
        conn.close()
        return group_id

    def delete_category_group(self, category_id: int) -> Dict[str, Any]:
        """Delete a category group if not used by rules"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        cursor.execute('SELECT major, minor FROM categories WHERE id = ?', (category_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return {'deleted': False, 'in_use': 0}

        category_name = self._format_category_name(row[0], row[1])
        cursor.execute('SELECT COUNT(*) FROM category_rules WHERE category = ?', (category_name,))
        in_use = cursor.fetchone()[0]

        if in_use > 0:
            conn.close()
            return {'deleted': False, 'in_use': in_use}

        cursor.execute('DELETE FROM categories WHERE id = ?', (category_id,))
        deleted = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return {'deleted': deleted, 'in_use': 0}

    def list_category_names(self, ledger_id: Optional[int] = None) -> List[str]:
        """Get full category names for dropdowns"""
        groups = self.get_category_groups(ledger_id)
        names = [self._format_category_name(group.major, group.minor) for group in groups]
        return [name for name in names if name]

    def category_group_conflict(self, major: str, minor: str, ledger_id: Optional[int], exclude_id: Optional[int] = None) -> bool:
        major = (major or "").strip()
        minor = (minor or "").strip()
        if not major:
            return False
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        query = "SELECT id FROM categories WHERE LOWER(major) = LOWER(?) AND LOWER(minor) = LOWER(?)"
        params: List[Any] = [major, minor]
        if ledger_id is not None:
            query += " AND (ledger_id IS NULL OR ledger_id = ?)"
            params.append(ledger_id)
        if exclude_id:
            query += " AND id != ?"
            params.append(exclude_id)
        cursor.execute(query, params)
        exists = cursor.fetchone() is not None
        conn.close()
        return exists

    def category_rule_combo_conflict(
        self,
        keyword: str,
        category_name: str,
        ledger_id: Optional[int],
        exclude_id: Optional[int] = None,
    ) -> bool:
        keyword = (keyword or "").strip()
        category_name = (category_name or "").strip()
        if not keyword or not category_name:
            return False
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        query = "SELECT id FROM category_rules WHERE LOWER(keyword) = LOWER(?) AND LOWER(category) = LOWER(?)"
        params: List[Any] = [keyword, category_name]
        if ledger_id is not None:
            query += " AND (ledger_id IS NULL OR ledger_id = ?)"
            params.append(ledger_id)
        if exclude_id:
            query += " AND id != ?"
            params.append(exclude_id)
        cursor.execute(query, params)
        exists = cursor.fetchone() is not None
        conn.close()
        return exists

    def _row_to_category_group(self, row) -> CategoryGroup:
        """Convert database row to CategoryGroup object"""
        return CategoryGroup(
            id=row[0],
            major=row[1],
            minor=row[2],
            created_at=row[3],
            updated_at=row[4],
            ledger_id=row[5] if len(row) > 5 else None
        )
    
    # Analytics methods
    def get_spending_summary(self, start_date: str = None, end_date: str = None,
                             keyword: str = None, major: str = None,
                             minor: str = None, ledger_id: Optional[int] = None,
                             include_in_budget: Optional[bool] = None) -> Dict[str, Any]:
        """Get spending summary for a date range"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()

        base_query = '''
            FROM bills b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE 1=1
        '''
        params = []

        if start_date:
            base_query += ' AND b.bill_date >= ?'
            params.append(start_date)

        if end_date:
            base_query += ' AND b.bill_date <= ?'
            params.append(end_date)

        categories_filter = self._fetch_category_names_by_major_minor(cursor, major, minor, ledger_id)
        base_query = self._append_category_filter(base_query, params, categories_filter, "COALESCE(c.major || '/' || c.minor, b.category)")

        if ledger_id is not None:
            base_query += ' AND b.ledger_id = ?'
            params.append(ledger_id)

        if include_in_budget is True:
            base_query += ' AND (b.include_in_budget = 1 OR b.include_in_budget IS NULL)'
        elif include_in_budget is False:
            base_query += ' AND b.include_in_budget = 0'

        if keyword:
            keyword_like = f'%{keyword}%'
            base_query += ' AND (b.merchant LIKE ? OR b.raw_text LIKE ?)'
            params.extend([keyword_like, keyword_like])

        cursor.execute(f'SELECT SUM(b.amount), COUNT(*) {base_query}', params)
        total_amount, bill_count = cursor.fetchone()

        cursor.execute(f'SELECT COUNT(DISTINCT b.bill_date) {base_query}', params)
        day_count = cursor.fetchone()[0] or 0

        # Get category breakdown
        category_query = f'''
            SELECT COALESCE(c.major || '/' || c.minor, b.category) AS cat_name,
                   SUM(b.amount), COUNT(*)
            {base_query}
            GROUP BY cat_name
            ORDER BY SUM(b.amount) DESC
        '''

        cursor.execute(category_query, params)
        categories = {}
        for row in cursor.fetchall():
            categories[row[0]] = {'amount': row[1], 'count': row[2]}

        conn.close()

        total_amount = total_amount or 0.0
        daily_avg = (total_amount / day_count) if day_count else 0.0

        return {
            'total_amount': total_amount,
            'bill_count': bill_count or 0,
            'categories': categories,
            'period_start': start_date,
            'period_end': end_date,
            'day_count': day_count,
            'daily_avg': daily_avg
        }
    
    def get_daily_spending(self, start_date: str = None, end_date: str = None,
                           keyword: str = None, major: str = None,
                           minor: str = None, ledger_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """Get daily spending data"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        query = '''
            SELECT b.bill_date, SUM(b.amount), COUNT(*)
            FROM bills b
            LEFT JOIN categories c ON b.category_id = c.id
            WHERE 1=1
        '''
        params = []
        
        if start_date:
            query += ' AND b.bill_date >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND b.bill_date <= ?'
            params.append(end_date)

        categories_filter = self._fetch_category_names_by_major_minor(cursor, major, minor, ledger_id)
        query = self._append_category_filter(query, params, categories_filter, "COALESCE(c.major || '/' || c.minor, b.category)")

        if ledger_id is not None:
            query += ' AND b.ledger_id = ?'
            params.append(ledger_id)

        if keyword:
            keyword_like = f'%{keyword}%'
            query += ' AND (b.merchant LIKE ? OR b.raw_text LIKE ?)'
            params.extend([keyword_like, keyword_like])
        
        query += ' GROUP BY b.bill_date ORDER BY b.bill_date DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        return [
            {
                'date': row[0],
                'amount': row[1],
                'count': row[2]
            }
            for row in rows
        ]


# Backward compatibility - Enhanced versions of existing classes
class EnhancedExcelSaver:
    """Enhanced Excel saver with date support"""
    
    def __init__(self, filename=config.EXCEL_PATH):
        self.filename = filename
        self.init_file()
    
    def init_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "账单记录"
            headers = ["记录时间", "账单日期", "截图文件名", "商户/商品", "分类", "金额", "备注(原始数据)"]
            ws.append(headers)
            wb.save(self.filename)
            print(f"📘 [Enhanced Excel] 已创建新账本: {self.filename}")
    
    def save(self, bill: EnhancedBill):
        """Save enhanced bill to Excel"""
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            
            merchant_clean = bill.merchant.split("￥")[0].replace(">", "").strip()
            row = [
                bill.created_at,
                bill.bill_date,
                bill.filename,
                merchant_clean,
                bill.category,
                bill.amount,
                str(bill.raw_text)[:50] + "..." if bill.raw_text else ""
            ]
            ws.append(row)
            wb.save(self.filename)
            print(f"✅ [Enhanced Excel] 写入成功")
        except PermissionError:
            print(f"❌ [Enhanced Excel] 写入失败: 请先关闭打开的文件！")


# Legacy compatibility wrapper
class DatabaseSaver:
    """Legacy compatibility wrapper for existing code"""
    
    def __init__(self, db_name=config.DB_PATH):
        self.enhanced_db = EnhancedDatabaseManager(db_name)
    
    def save(self, data, image_name, bill_date=None):
        """Save bill data (legacy format)"""
        bill = EnhancedBill(
            filename=image_name,
            merchant=data.get("merchant", ""),
            amount=data.get("amount", 0.0),
            category=data.get("category", ""),
            raw_text=data.get("raw_text", []),
            bill_date=bill_date or datetime.date.today().strftime("%Y-%m-%d"),
            is_manual=False
        )
        return self.enhanced_db.save_bill(bill)
